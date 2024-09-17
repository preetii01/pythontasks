# TASK-2: CREATE 3 TYPES OF FILES:
'''1.JSON FILE
   2.CSV FILE
   3.EXCEL FILE   
   (Read and write your data-name, email, college name in all three files)
   '''

#JSON FILE

#WRITE
import json

dic={"name":"preeti","college_name":"ITS ghaziabad","email":"bhardwajpreeti2102@gmail.com"}
json_obj=json.dumps(dic)  #dumps convert a subset of python objects into a json string 
file=open("text.json","w")
file.write(json_obj)
file.close()

#READ
file=open("text.json","r")
print(file.read())

#CSV FILE

#WRITE
import csv
header=['name','college name','email']
data=['preeti','ITS ghaziabad','bhardwajpreeti2102@gmail.com']
filename='data.csv'
with open ('data.csv','w',newline="") as file:
   csvwriter=csv.writer(file)
   csvwriter.writerow(header)
   csvwriter.writerows(data)
   
#READ
with open('data.csv','r') as csvfile:
   csv_reader=csv.reader(csvfile)
   for row in csv_reader:
      print(row)
   
#EXCEL

#WRITE
import openpyxl
wb=openpyxl.Workbook()
ws=wb.active
ws["A1"]="NAME"
ws["B1"]="COLLEGE NAME"
ws["C1"]="EMAIL"
ws["A2"]="Preeti"
ws["B2"]="ITS ghaziabad"
ws["C2"]="bhardwajpreeti2102@gmail.com"
wb.save("details.xlsx")

#READ
wb=openpyxl.load_workbook("details.xlsx")
ws=wb.active
for row in range(0,ws.max_row):
   for col in ws.iter_cols(1,ws.max_column):
      print(col[row].value)




